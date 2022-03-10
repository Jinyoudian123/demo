# coding=utf8
import openpyxl
import pymysql
import yaml
import zmail


def read_excel(filepath):
    """
        excel文件读取的方法 \n
        返回字典数据类型，key是sheet名，value为该sheet下的所有内容
    """
    sheets = openpyxl.load_workbook(filepath)
    sheetnames = sheets.sheetnames
    excel_data = {}
    # 遍历出所有sheet的名称
    for sheet in sheetnames:
        row_datas = []
        # 遍历出该sheet下每行的数据
        for row_data in sheets[sheet].values:
            if row_data[0] is not None:
                row_datas.append(row_data)
        # 只将有数据的sheet添加到excel_data中
        if row_datas != []:
            excel_data[sheet] = row_datas
    return excel_data


def read_yaml(filepath):
    with open(filepath, 'r', encoding='utf-8') as file:
        yaml_data = yaml.load(file, Loader=yaml.FullLoader)
        return yaml_data


def sql_select(sql):
    # 打开数据库连接
    db = pymysql.connect(host="localhost", user="root",
                         password="root", db="demo1", port=3306)
    # 使用cursor()方法获取操作游标
    cur = db.cursor()
    # 1.查询操作
    # 编写sql 查询语句  user 对应我的表名
    # sql = "SELECT * FROM goods"
    # sql = "SELECT * FROM goods WHERE tid=0"
    try:
        cur.execute(sql)  # 执行sql语句
        results = cur.fetchall()  # 获取查询的所有记录
        # print("id", "name", "password")
        # 遍历结果
        # print('results:',results)
        if len(results) == 0:
            return None
        if len(results) == 1:
            if len(results[0]) == 1:
                return results[0][0]
            return results[0]
        return results
    except Exception as e:
        return 'sql语句有误，请检查', e
    finally:
        db.close()  # 关闭连接


def send_email(sender_user, sender_pwd, recevier_list, report_url):
    """
    发送邮件方法
    :param sender_user: 发送人账号
    :param sender_pwd: 发送人密码
    :param recevier_list: 接收人账号，list类型
    :param report_url: 测试报告的地址，str类型
    :return:
    """
    mail_content = {
        "Subject": "接口测试报告",
        "Content_html": f'<a href="{report_url}">点击查看测试报告</a>',
    }
    # 发送人
    sender = {'username': sender_user, 'pwd': sender_pwd}
    # 收件人
    recevier = recevier_list
    # 登录邮箱
    server = zmail.server(sender['username'], sender['pwd'])
    # 服务器发送邮件
    server.send_mail(recevier, mail_content)


if __name__ == '__main__':
    # result = sql_select('SELECT goods_id,goods_name,tid FROM goods WHERE goods_id=1')
    result = sql_select('SELECT goods_name,tid FROM goods WHERE goods_id=1')
    print(type(result))
    print(len(result))
    print(result)
    if result == ((1, 'Air笔记本电脑', 1),):
        print(12)
